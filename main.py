# импортируем необходимые библиотеки
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# чтение файла Источник.XLS и создание временной таблицы istochnik
istochnik = pd.read_excel("Источник.XLS", usecols=[0, 1, 3, 4, 6], header=None,
                          names=["modul", "kanal", "tipe", "signal", "koment"])

# замена символов в колонке signal
istochnik["signal"] = istochnik["signal"].str.replace("ШВП", "SHVP").str.replace("е", "e").str.replace("с", "c").str.replace("о", "o").str.replace("р", "p").str.replace(".", "_")



# получение уникальных значений из столбца modul
modul_values = istochnik["modul"].unique()

# запись таблиц в файл Результат.xlsx
with pd.ExcelWriter("Результат.xlsx") as writer:
    for modul in modul_values:
        # фильтрация таблицы istochnik по значению modul и удаление столбца modul
        filtered_table = istochnik.loc[istochnik["modul"] == modul, ["kanal", "tipe", "signal", "koment"]]
        # запись таблицы в файл Результат.xlsx с именем листа modul
        filtered_table.to_excel(writer, sheet_name=modul, header=False, index=False)
        # получение последней записи (ячейки) в таблице и ее координат
        last_cell = writer.sheets[modul].cell(row=writer.sheets[modul].max_row, column=writer.sheets[modul].max_column)

        # подбор ширины колонок по содержимому
        for col in range(1, last_cell.column + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in writer.sheets[modul][column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            writer.sheets[modul].column_dimensions[column].width = adjusted_width

        # создание стиля для обводки таблицы жирной линией
        border_style = Border(left=Side(style='thick'),
                              right=Side(style='thick'),
                              top=Side(style='thick'),
                              bottom=Side(style='thick'))
        # получение ячеек, которые нужно обвести линией
        cells_range = writer.sheets[modul]["A1":last_cell.coordinate]
        # применение стиля ко всем ячейкам в таблице
        for row in cells_range:
            for cell in row:
                cell.border = border_style
        #cells_range.border = border_style


    # Разбиваем временную таблицу istochnik на несколько таблиц по критерию modul
    tables = {}
    for modul in istochnik['modul'].unique():
        tables[modul] = istochnik[istochnik['modul'] == modul].drop(columns=['modul'])




    # создание последнего листа
    wb = writer.book
    last_sheet = wb.create_sheet("Итог")

    # Записываем таблицы на один лист
    for i, (modul, table) in enumerate(tables.items()):
        table.to_excel(writer, header=False, sheet_name="Итог", startrow=2, startcol=i * table.shape[1]+2*i+2, index=False)
        last_sheet.append([])  # добавление пустой строки между таблицами

    # подбор ширины колонок по содержимому
    for col in range(1, last_sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in last_sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        last_sheet.column_dimensions[column].width = adjusted_width



print("Готово!")